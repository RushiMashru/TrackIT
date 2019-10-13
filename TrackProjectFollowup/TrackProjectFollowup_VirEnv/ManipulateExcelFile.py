import openpyxl
import os
from openpyxl import load_workbook
import TrackFollowUpModel as TM

EmailSent = 'Yes'
RespondedYes = 'Yes'
dirPath = os.path.dirname(os.path.realpath('__file__'))
path = os.path.join(dirPath, 'Test_TrackFollowup.xlsx')

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
sheet_obj.title = 'TrackIT'


def GetAllRowsByColumn(col_num):
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=col_num)
        print(cell_obj.value)


def GetAllColumnsByRow(row_num):
    T = []
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column

    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=row_num, column=i)
        T.append(cell_obj.value)

    oTrackIt = TM.TrackFollowUpModel(T[0], T[1], T[2], T[3], T[4], T[5], T[6], T[7],
                                     T[8], T[9], T[10], T[11], T[12], T[13], T[14], T[15], T[16], T[17], T[18], T[19], T[20])
    return oTrackIt


def GetSpecificValue(row_num, col_num):
    cell_obj = sheet_obj.cell(row=row_num, column=col_num)
    print(cell_obj.value)


def GetSheetTitle():
    print('Sheet name is', sheet_obj.title)


def GetMaxRow():
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    return max_row


def GetMaxCol():
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    return max_col


def GetListFromExcel():
    DataList = []
    T = []
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            T.append(cell_obj.value)

        DataList.append(TM.TrackFollowUpModel(
            T[0], T[1], T[2], T[3], T[4], T[5], T[6], T[7], T[8], T[9], T[10], T[11], T[12], T[13], T[14], T[15], T[16], T[17], T[18], T[19], T[20]))
        T.clear()

    return DataList


def UpdateEmailSentBySystemStatus(row_n, col_n):
    oCell = sheet_obj.cell(row=row_n, column=col_n)
    oCell.value = EmailSent
    wb_obj.save(path)


def UpdateResponseFromReceiverStatus(row_n, col_n, response):
    oCell = sheet_obj.cell(row=row_n, column=col_n)
    oCell.value = response
    wb_obj.save(path)


def AddNewEntryToExcel(dataList=[]):
    if len(dataList) > 0:
        skipColumns = [13, 16, 17, 18, 19, 20]
        for index, item in enumerate(dataList):
            if(int(index)+1 in skipColumns):
                continue
            oCell = sheet_obj.cell(row=dataList[0], column=int(index)+1)
            oCell.value = item

        wb_obj.save(path)


def UpdateFollowUpDate(row_n, col_n, updatedDate):
    oCell = sheet_obj.cell(row=row_n, column=col_n)
    oCell.value = updatedDate
    wb_obj.save(path)
